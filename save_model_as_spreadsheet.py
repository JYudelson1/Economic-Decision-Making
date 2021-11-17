import pickle as pkl
from openpyxl import Workbook #type: ignore
from openpyxl.styles import Font, Alignment #type: ignore
from utils import *
from ev_based_model import EVModel
from pt_predictor import PTModel
from hpt_predictor import HPTTWModel
from eut_predictor import EUTModel

fullnames = {
    "a": "alpha",
    "b": "beta",
    "g": "gamma",
    "l": "lambda",
    "tw": "tw",
    "xl": "chi-loss",
    "xg": "chi-gain"
}

def spreadsheet_main(version: str, filename: str, use_all_fits: bool = True) -> None:

    # Load model
    with open(version, "rb") as f:
        model = pkl.load(f)

    workbook = Workbook()

    # Make sheets
    paramsheet = workbook.active
    paramsheet.title = "Parameters"
    predsheet = workbook.create_sheet("Predictions")

    # Add headers to first sheet
    paramsheet["A1"] = "Subject"
    err_col = 2
    for i, cell in enumerate(paramsheet.iter_cols(min_col = 2,
                                                  max_col = 6,
                                                  min_row = 1,
                                                  max_row = 1)):
        if i >= len(model.free_params):
            break
        cell[0].value = fullnames[model.free_params[i]] + "New"
        err_col += 1
    paramsheet.cell(row=1, column=err_col, value="ErrorsNew")
    paramsheet.cell(row=1, column=err_col + 1, value="ErrorsOld")
    for i, cell in enumerate(paramsheet.iter_cols(min_col = err_col + 2,
                                                  max_col = err_col + 5,
                                                  min_row = 1,
                                                  max_row = 1)):
        if i >= len(model.free_params):
            break
        cell[0].value = fullnames[model.free_params[i]] + "Old"

    # Add headers to second sheet
    predsheet["A1"] = "Subject"
    predsheet["B1"] = "Day"
    for day in range(68):
        predsheet.merge_cells(start_row=1, start_column=(day*4+3), end_row=1, end_column=(day*4+6))
        merge = predsheet.cell(row=1, column=(day*4+3), value=(67 - day))
        merge.alignment = Alignment(horizontal='center')

        predsheet.cell(row=2, column=(day*4+3), value="Stored")
        predsheet.cell(row=2, column=(day*4+4), value="Price")
        predsheet.cell(row=2, column=(day*4+5), value="Sold")
        predsheet.cell(row=2, column=(day*4+6), value="Predicted")


    # Bold the first row of everything
    bold = Font(bold=True)
    for cell in paramsheet["1:1"]:
        cell.font = bold
    for r in paramsheet["1:2"]:
        for cell in r:
            cell.font = bold


    # Iterate through each patient, getting each best fit
    current_row_pred = 3
    current_row_param = 3
    for s in trange(57, desc="Saving..."):
        subject = s + 1

        ## Save prediction data & best fit data
        fit = model.best_fits[s]
        paramsheet[f'A{current_row_param}'] = subject
        predsheet[f'A{current_row_param}'] = subject

        # Get predictions for that fit
        predictions: List[int] = model.predict_one_subject(s, fit)

        # Get and save errors
        error = model.mean_error_one_subject_proportion(s, predictions)

        paramsheet.cell(row=current_row_param, column=err_col, value=round(error, 3))

        if use_all_fits:
            all_fits = model.all_best_fits[s]
            for good_fit in all_fits:
                # save only the free params
                for i, param in enumerate(good_fit.free_params):
                    # Rounding just to remove floating point error
                    paramsheet[current_row_param][i+1].value = str(round(getattr(good_fit,param), 3))
                current_row_param += 1
        else:
            for i, param in enumerate(fit.free_params):
                # Rounding just to remove floating point error
                paramsheet[current_row_param][i+1].value = str(round(getattr(fit,param), 3))
            current_row_param += 1

        # Iterate through each day
        for d in range(NUM_DAYS):
            day = NUM_DAYS - d

            # Save info
            predsheet.cell(row=current_row_pred, column=(day*4-1), value=int(model.data.loc[s]['stored'][d]))
            predsheet.cell(row=current_row_pred, column=(day*4), value=int(model.data.loc[s]['price'][d]))
            predsheet.cell(row=current_row_pred, column=(day*4+1), value=int(model.data.loc[s]['sold'][d]))
            predsheet.cell(row=current_row_pred, column=(day*4+2), value=predictions[d])

        current_row_pred += 1

    # Save final mean error & std dev of error
    # all_errors: List[float] = model.mean_error_all_subjects(verbose=False)
    # errorsheet[f'A{current_row + 1}'] = "Mean Error:"
    # errorsheet[f'B{current_row + 1}'] = np.mean(all_errors)
    # errorsheet[f'A{current_row + 2}'] = "Mean Percent Error:"
    # errorsheet[f'B{current_row + 2}'] = str(100 * np.mean(all_errors)) + "%"
    # errorsheet[f'A{current_row + 3}'] = "Standard Deviation:"
    # errorsheet[f'B{current_row + 3}'] = np.std(all_errors)

    # Adjust size of errorsheet to fit:
    # cell_to_text = lambda x: str(x) if x else ""
    # for col in errorsheet.columns:
    #     length = max(len(cell_to_text(cell)) for cell in col)
    #     errorsheet.column_dimensions[col[0].column_letter].width = length

    workbook.save(filename=filename)

if __name__ == '__main__':

    filename = "data/pt_predictions_full_iter_1029.xlsx"
    version = "data/v2_exhaustive_iter_full_1029.pkl"
    spreadsheet_main(version=version, filename=filename)
